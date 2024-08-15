from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from locators import Locators

class ExcelAction:
    def __init__(self):
        pass

    def write_match_details(self, match_details_list):
        self.workbook = Workbook()
        for sheet in Locators.excel_sheets:
            self.workbook.create_sheet(title=sheet)
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        print("Workbook created successfully.")
        scorecard_sheet = self.workbook['Scorecard']
        for index in range(len(match_details_list)):
            current_cell = scorecard_sheet.cell(row=index+1, column=1, value=match_details_list[index])
            current_cell.font = Font(bold=True)      
        file_name = match_details_list[0]
        self.sanitized_file_name = file_name.replace(",", "").replace("\n", "").replace(" ", "_").replace("/","")
        self.workbook.save(f"{self.sanitized_file_name}.xlsx")
        return True

    def write_match_stats(self, bat_stat_list, sheetname):
        self.bat_stat_list = bat_stat_list
        self.sheetname = sheetname
        current_sheet = self.workbook[sheetname]
        last_used_row = self.get_last_used_row(sheetname)
        # Change the background color to yellow
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        count = 0
        for bat_stat in bat_stat_list:
            if count == 0: 
                for i in range(len(bat_stat)):
                    current_cell = current_sheet.cell(row=last_used_row+1, column=i+1, value=bat_stat[i])
                    current_cell.font = Font(bold=True)
                    current_cell.fill = yellow_fill
                last_used_row = self.get_last_used_row(sheetname)
            else:
                for i in range(len(bat_stat)):
                    current_sheet.cell(row=last_used_row+i+1, column=count, value=bat_stat[i])
            count += 1
            self.workbook.save(f"{self.sanitized_file_name}.xlsx")
            
    def get_last_used_row(self, sheetname):
        self.sheetname = sheetname
        current_sheet = self.workbook[sheetname]
        # Iterate from the last row to the first row
        for row in reversed(range(1, current_sheet.max_row + 1)):
            for col in current_sheet.iter_cols(1, current_sheet.max_column):
                if col[row-1].value is not None:
                    return row
                else:
                    return 0
